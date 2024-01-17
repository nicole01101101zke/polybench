"""
    aggregateOperationInstCount: extractvalue insertvalue
    arithmeticInstCount: add sub mul udiv sdiv urem srem icmp
    bitwiseOperationInstCount: shl lshr ashr and or xor
    callFunctionInstCount: call
    conversionInstCount: trunc zext sext fptrunc fpext fptoui fptosi uitofp sitofp ptrtoint inttoptr bitcast addrspacecast
    floatOperationInstCount: fadd fsub fmul fdiv frem fneg fcmp
    memoryAccessInstCount: alloca load store fence cmpxchg atomicrmw getelementptr
    terminatorInstCount: ret br switch indirectbr invoke callbr resume catchswitch catchret cleanupret unreachable
    vectorOperationInstCount: extractelement insertelement shufflevector
"""
agstrings_to_check = ["extractvalue", "insertvalue"]
arstrings_to_check = ["add", "sub", "mul", "udiv", "sdiv", "urem", "srem", "icmp"]
bstrings_to_check = ["shl", "lshr", "ashr", "and", "or", "xor"]
castrings_to_check = ["call"]
costrings_to_check = ["trunc", "zext", "sext", "fptrunc", "fpext", "fptoui", "fptosi", "uitofp", "sitofp", "ptrtoint", "inttoptr", "bitcast", "addrspacecast"]
fstrings_to_check = ["fadd", "fsub", "fmul", "fdiv", "frem", "fneg", "fcmp"]
mstrings_to_check = ["alloca", "load", "store", "fence", "cmpxchg", "atomicrmw", "getelementptr"]
tstrings_to_check = ["ret", "br", "switch", "indirectbr", "invoke", "callbr", "resume", "catchswitch", "catchret", "cleanupret", "unreachable"]
vstrings_to_check = ["extractelement", "insertelement", "shufflevector"]

import openpyxl

def check_strings(main_string, *strings_to_check):
    for string in strings_to_check:
        if string in main_string:
            return True
    return False

def count_instructions(file_path, benchmark_row0):
    aggregateOperationInstCount = 0
    arithmeticInstCount = 0
    bitwiseOperationInstCount = 0
    callFunctionInstCount = 0
    conversionInstCount = 0
    floatOperationInstCount = 0
    memoryAccessInstCount = 0
    terminatorInstCount = 0
    vectorOperationInstCount = 0

    # 打开.ll文件进行读取
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # 遍历文件中的每一行
    for line in lines:
        if line.startswith(';'):
            continue
        
        if check_strings(line, *agstrings_to_check):
            aggregateOperationInstCount += 1

        if check_strings(line, *arstrings_to_check):
            arithmeticInstCount += 1

        if check_strings(line, *bstrings_to_check):
            bitwiseOperationInstCount += 1

        if check_strings(line, *castrings_to_check):
            callFunctionInstCount += 1

        if check_strings(line, *costrings_to_check):
            conversionInstCount += 1

        if check_strings(line, *fstrings_to_check):
            floatOperationInstCount += 1

        if check_strings(line, *mstrings_to_check):
           memoryAccessInstCount += 1

        if check_strings(line, *tstrings_to_check):
            terminatorInstCount += 1

        if check_strings(line, *vstrings_to_check):
            vectorOperationInstCount += 1


    # 打印统计结果
    print(f"聚合指令数量: {aggregateOperationInstCount}")
    print(f"算数指令数量: {arithmeticInstCount}")
    print(f"位运算指令数量: {bitwiseOperationInstCount}")
    print(f"调用指令数量: {callFunctionInstCount}")
    print(f"转换指令数量: {conversionInstCount}")
    print(f"浮点指令数量: {floatOperationInstCount}")
    print(f"访存指令数量: {memoryAccessInstCount}")
    print(f"终止指令数量: {terminatorInstCount}")
    print(f"向量指令数量: {vectorOperationInstCount}")

    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook.active
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row[0] and row[0].startswith(benchmark_row0):
            sheet.cell(row=row_number, column=2, value=aggregateOperationInstCount)
            sheet.cell(row=row_number, column=3, value=arithmeticInstCount)
            sheet.cell(row=row_number, column=4, value=bitwiseOperationInstCount)
            sheet.cell(row=row_number, column=5, value=callFunctionInstCount)
            sheet.cell(row=row_number, column=6, value=conversionInstCount)
            sheet.cell(row=row_number, column=7, value=floatOperationInstCount)
            sheet.cell(row=row_number, column=8, value=memoryAccessInstCount)
            sheet.cell(row=row_number, column=9, value=terminatorInstCount)
            sheet.cell(row=row_number, column=10, value=vectorOperationInstCount)
            break
    workbook.save('data.xlsx')


    
# 调用函数并传入.ll文件的路径
count_instructions('../datamining/correlation/correlation.ll', 'polybench_correlation')
count_instructions('../datamining/covariance/covariance.ll', 'polybench_covariance')
count_instructions('../linear-algebra/kernels/2mm/2mm.ll', 'polybench_2mm')
count_instructions('../linear-algebra/kernels/3mm/3mm.ll', 'polybench_3mm')
count_instructions('../linear-algebra/kernels/atax/atax.ll', 'polybench_atax')
count_instructions('../linear-algebra/kernels/bicg/bicg.ll', 'polybench_bicg')
count_instructions('../linear-algebra/kernels/doitgen/doitgen.ll', 'polybench_doitgen')
count_instructions('../linear-algebra/kernels/gemm/gemm.ll', 'polybench_gemm')
count_instructions('../linear-algebra/kernels/gemver/gemver.ll', 'polybench_gemver')
count_instructions('../linear-algebra/kernels/gesummv/gesummv.ll', 'polybench_gesummv')
count_instructions('../linear-algebra/kernels/mvt/mvt.ll', 'polybench_mvt')
count_instructions('../linear-algebra/kernels/syr2k/syr2k.ll', 'polybench_syr2k')
count_instructions('../linear-algebra/kernels/syrk/syrk.ll', 'polybench_syrk')
count_instructions('../linear-algebra/solvers/gramschmidt/gramschmidt.ll', 'polybench_gramschmidt')
count_instructions('../linear-algebra/solvers/lu/lu.ll', 'polybench_lu')
count_instructions('../stencils/adi/adi.ll', 'polybench_adi')
count_instructions('../stencils/convolution-2d/2DConvolution.ll', 'polybench_convolution-2d')
count_instructions('../stencils/convolution-3d/3DConvolution.ll', 'polybench_convolution-3d')
count_instructions('../stencils/fdtd-2d/fdtd2d.ll', 'polybench_fdtd-2d')
count_instructions('../stencils/jacobi-1d-imper/jacobi1D.ll', 'polybench_jacobi-1d-imper')
count_instructions('../stencils/jacobi-2d-imper/jacobi2D.ll', 'polybench_jacobi-2d-imper')
